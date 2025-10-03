from __future__ import annotations

import io
from collections.abc import Iterable
from typing import Any, Dict, List, Mapping

from openpyxl import load_workbook

__all__ = ["parse_sheet"]


def parse_sheet(
    source: Any,
    *,
    sheet_name: str | None = None,
    header_row: int = 1,
    trim_headers: bool = True,
) -> List[Mapping[str, Any]]:
    """Parse an Excel sheet into a list of dictionaries."""

    workbook = _load_workbook(source)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return []

    header_index = max(header_row - 1, 0)
    headers_row = rows[header_index]
    headers = _normalise_headers(headers_row, trim=trim_headers)

    data_rows = rows[header_index + 1 :]
    parsed: List[Mapping[str, Any]] = []
    for row in data_rows:
        if row is None:
            continue
        values = list(row)
        if all(value is None or value == "" for value in values):
            continue
        record: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if header is None:
                continue
            record[header] = values[idx] if idx < len(values) else None
        parsed.append(record)

    return parsed


def _load_workbook(source: Any):
    if isinstance(source, (bytes, bytearray)):
        return load_workbook(io.BytesIO(source), read_only=True, data_only=True)
    if hasattr(source, "read"):
        data = source.read()
        if isinstance(data, (bytes, bytearray)):
            return load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        raise TypeError("Excel source must provide bytes-like data")
    if isinstance(source, str):
        return load_workbook(filename=source, read_only=True, data_only=True)
    raise TypeError("Unsupported Excel source type")


def _normalise_headers(headers: Iterable[Any], *, trim: bool) -> List[str | None]:
    normalised: List[str | None] = []
    for index, header in enumerate(headers, start=1):
        if header is None:
            normalised.append(f"column_{index}")
            continue
        header_text = str(header)
        normalised.append(header_text.strip() if trim else header_text)
    return normalised
